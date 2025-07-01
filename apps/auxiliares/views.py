import os
from django.conf                        import settings
import openpyxl
from django.http                        import HttpResponse
from io                                 import BytesIO
from apps.cuenta.models                 import User        # Importa el modelo User si aún no lo tienes
from datetime                           import date
from apps.gestion.models.director       import Director
from apps.gestion.models.asignacion     import Asignacion
from apps.auxiliares.models.responsable import Responsable

# from configuracion  import settings

def reporte_bienes(request, id_responsable):

    ruta_plantilla = os.path.join(settings.MEDIA_ROOT, 'plantillas', 'formato_inventario.xlsx')
    if not os.path.exists(ruta_plantilla):
        return HttpResponse("Plantilla no encontrada.", status=404)

    wb = openpyxl.load_workbook(ruta_plantilla)
    ws = wb.active

    # Obtener bienes del responsable desde el ORM
    bienes      = Asignacion.objects.filter(responsable_id = id_responsable)
    responsable = Responsable.objects.filter(id=id_responsable).first()
    # director = Director.objects.filter(responsable_id  =  id_responsable)
    # Obtener los datos del responsable del levantamiento del inventario
    usuario =request.user


    fila_inicio = 30  # Asumiendo que empiezas a escribir desde la fila 30
    for i, bien in enumerate(bienes, start=0):
        fila = fila_inicio + i
        # ws[f'A{fila}'] = bien.cantidad
        ws[f'B{fila}'] = f'{bien.bien.categoria.cod_catalogo}'
        ws[f'C{fila}'] = bien.bien.cod_bien
        ws[f'D{fila}'] = f'{bien.bien.categoria} {bien.bien.modelo}'
        ws[f'E{fila}'] = bien.bien.valor_unitario
        ws[f'F{fila}'] = bien.bien.valor_unitario


    #Usuario encargado del levantamiento del inventario
    ws['B43'] = f'{usuario.origen}-{usuario.cedula}'
    ws['D43'] = usuario.nombre_apellido 

    #Datos del  director
    # ws['B45'] = f'{usuario.}'

    #Unidad usuaria
    ws['B26'] = f'DENOMINACIÓN: {bien.responsable.unidad}'
    fecha_hoy = date.today()
    fecha_formateada = fecha_hoy.strftime('%d/%m/%Y')
    ws['F17'] = f'FECHA: {fecha_formateada}'

    #Almacen responsable
    #Cédula 
    ws['C28'] = f'C.I. {responsable.persona.origen}-{responsable.persona.cedula}'
    #Nombres apellidos
    ws['D28'] = f'APELLIDOS Y NOMBRES:  {responsable.persona.nombres_apellidos}'
    #Cargo
    ws['E28'] = f'CARGO:  {responsable.persona.cargo}'


   
    

    # Guardar en memoria y devolver
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="reporte_bienes_{id_responsable}.xlsx"'
    return response
