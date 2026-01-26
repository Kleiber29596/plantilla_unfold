import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from apps.bien.models.detalle_asignacion import DetalleAsignacion
from apps.bien.models.asignaciones import Asignacion
from apps.auxiliares.models.dependencia import Dependencia
from apps.auxiliares.models.subdependencia import Subdependencia
from datetime import datetime
import re
import logging

# Configurar logging para debug
logger = logging.getLogger(__name__)

# --- Funciones de Estilo ---

def get_estilos():
    """Retorna un diccionario con todos los estilos visuales"""
    
    COLOR_PRIMARIO = "B7B7B7"
    COLOR_FONDO_CLARO = "F8FAFC"
    COLOR_TEXTO_OSCURO = "1F2937"
    COLOR_BORDE = "CBD5E1"
    COLOR_ENCABEZADO = "B7B7B7"
    
    borde_delgado = Border(
        left=Side(style='thin', color=COLOR_BORDE),
        right=Side(style='thin', color=COLOR_BORDE),
        top=Side(style='thin', color=COLOR_BORDE),
        bottom=Side(style='thin', color=COLOR_BORDE)
    )
    
    borde_grueso = Border(
        left=Side(style='medium', color=COLOR_PRIMARIO),
        right=Side(style='medium', color=COLOR_PRIMARIO),
        top=Side(style='medium', color=COLOR_PRIMARIO),
        bottom=Side(style='medium', color=COLOR_PRIMARIO)
    )
    
    # Crear objetos PatternFill explícitamente
    fill_primario = PatternFill(start_color=COLOR_PRIMARIO, end_color=COLOR_PRIMARIO, fill_type="solid")
    fill_encabezado = PatternFill(start_color=COLOR_ENCABEZADO, end_color=COLOR_ENCABEZADO, fill_type="solid")
    fill_fondo_claro = PatternFill(start_color=COLOR_FONDO_CLARO, end_color=COLOR_FONDO_CLARO, fill_type="solid")
    fill_vacio = PatternFill(fill_type=None)
    
    return {
        'titulo_principal': {
            'font': Font(name='Calibri', size=16, bold=True, color=COLOR_TEXTO_OSCURO),
            'fill': fill_primario,
            'alignment': Alignment(horizontal="center", vertical="center", wrap_text=True),
            'border': borde_grueso
        },
        'titulo_responsable': {
            'font': Font(name='Calibri', size=14, bold=True, color=COLOR_TEXTO_OSCURO),
            'fill': fill_primario,
            'alignment': Alignment(horizontal="center", vertical="center", wrap_text=True),
            'border': borde_grueso
        },
        'encabezado_tabla': {
            'font': Font(name='Calibri', size=11, bold=True, color=COLOR_TEXTO_OSCURO),
            'fill': fill_encabezado,
            'alignment': Alignment(horizontal="center", vertical="center", wrap_text=True),
            'border': borde_delgado
        },
        'celda_datos': {
            'font': Font(name='Calibri', size=10, color=COLOR_TEXTO_OSCURO),
            'fill': fill_vacio,
            'alignment': Alignment(horizontal="center", vertical="center"),
            'border': borde_delgado
        },
        'celda_datos_alterna': {
            'font': Font(name='Calibri', size=10, color=COLOR_TEXTO_OSCURO),
            'fill': fill_fondo_claro,
            'alignment': Alignment(horizontal="center", vertical="center"),
            'border': borde_delgado
        },
        'celda_izquierda': {
            'font': Font(name='Calibri', size=10, color=COLOR_TEXTO_OSCURO),
            'fill': fill_vacio,
            'alignment': Alignment(horizontal="left", vertical="center", wrap_text=True),
            'border': borde_delgado
        },
        'celda_izquierda_alterna': {
            'font': Font(name='Calibri', size=10, color=COLOR_TEXTO_OSCURO),
            'fill': fill_fondo_claro,
            'alignment': Alignment(horizontal="left", vertical="center", wrap_text=True),
            'border': borde_delgado
        },
        'etiqueta': {
            'font': Font(name='Calibri', size=11, bold=True, color=COLOR_TEXTO_OSCURO),
            'fill': fill_vacio,
            'alignment': Alignment(horizontal="left", vertical="center"),
            'border': borde_delgado
        },
        'info_valor': {
            'font': Font(name='Calibri', size=11, color=COLOR_TEXTO_OSCURO),
            'fill': fill_vacio,
            'alignment': Alignment(horizontal="left", vertical="center"),
            'border': borde_delgado
        },
        'total': {
            'font': Font(name='Calibri', size=11, bold=True, color=COLOR_TEXTO_OSCURO),
            'fill': fill_encabezado,
            'alignment': Alignment(horizontal="center", vertical="center"),
            'border': borde_delgado
        }
    }

def aplicar_estilo_celda(celda, estilo):
    """Aplica un estilo completo a una celda de manera segura"""
    if estilo:
        if 'font' in estilo and estilo['font']:
            celda.font = estilo['font']
        if 'fill' in estilo and estilo['fill']:
            celda.fill = estilo['fill']
        if 'alignment' in estilo and estilo['alignment']:
            celda.alignment = estilo['alignment']
        if 'border' in estilo and estilo['border']:
            celda.border = estilo['border']

def sanitizar_nombre_hoja(nombre):
    """Limpia el nombre para que sea válido como nombre de hoja en Excel"""
    if not nombre:
        return "SinNombre"
    
    nombre = str(nombre)
    # Reemplazar caracteres inválidos
    nombre = re.sub(r'[\\/*?:[\]]', '_', nombre)
    # Reemplazar espacios por guiones bajos
    nombre = nombre.replace(' ', '_')
    # Limitar longitud
    if len(nombre) > 30:
        nombre = nombre[:27] + "..."
    return nombre

def reordenar_hojas(wb):
    """Reordena las hojas en un orden lógico - VERSIÓN SIMPLIFICADA"""
    # Orden deseado
    orden_deseado = ["Índice", "Estadísticas"]
    
    # Obtener todas las hojas
    sheets = wb._sheets
    
    # Crear diccionario de hojas por nombre
    hojas_por_nombre = {sheet.title: sheet for sheet in sheets}
    
    # Crear nueva lista de hojas en orden deseado
    nuevo_orden = []
    
    # Primero agregar las hojas en el orden deseado
    for nombre in orden_deseado:
        if nombre in hojas_por_nombre:
            nuevo_orden.append(hojas_por_nombre[nombre])
    
    # Luego agregar el resto de hojas (hojas de responsables)
    for sheet in sheets:
        if sheet.title not in orden_deseado:
            nuevo_orden.append(sheet)
    
    # Reemplazar la lista de hojas
    wb._sheets = nuevo_orden

# --- FUNCIONES AUXILIARES (DEBEN ESTAR ANTES DE LA FUNCIÓN PRINCIPAL) ---

def crear_reporte_error(response, mensaje):
    """Crea un reporte de error cuando hay problemas"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Error"
    
    ws['A1'] = "ERROR AL GENERAR REPORTE"
    ws['A1'].font = Font(name='Calibri', size=14, bold=True, color="FF0000")
    
    ws['A3'] = mensaje
    ws['A3'].font = Font(name='Calibri', size=12)
    
    ws['A5'] = "Fecha:"
    ws['B5'] = datetime.now().strftime("%d/%m/%Y %H:%M")
    
    # Ajustar anchos
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 30
    
    wb.save(response)
    return response

def crear_reporte_subdependencia_vacia(response, subdependencia):
    """Crea un reporte cuando la subdependencia no tiene bienes"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Información"
    
    estilos = get_estilos()
    
    fecha_actual = datetime.now().strftime("%d/%m/%Y")
    
    # Título
    ws.merge_cells('A1:G1')
    ws['A1'] = f"INVENTARIO DE BIENES - {subdependencia.nombre.upper()}"
    aplicar_estilo_celda(ws['A1'], estilos['titulo_principal'])
    ws.row_dimensions[1].height = 35
    
    # Fecha
    ws['A2'] = f"Fecha: {fecha_actual}"
    ws['A2'].font = Font(name='Calibri', size=10, italic=True)
    
    # Información
    ws['A4'] = "DEPENDENCIA:"
    ws['B4'] = subdependencia.dependencia.nombre
    
    ws['A5'] = "SUBDEPENDENCIA:"
    ws['B5'] = subdependencia.nombre
    
    # Mensaje
    ws.merge_cells('A7:G7')
    ws['A7'] = "NO SE ENCONTRARON BIENES ASIGNADOS EN ESTA SUBDEPENDENCIA"
    ws['A7'].font = Font(name='Calibri', size=12, bold=True, color="FF0000")
    ws['A7'].alignment = Alignment(horizontal="center", vertical="center")
    
    # Información de debug
    ws['A9'] = "Información para debug:"
    ws['A9'].font = Font(name='Calibri', size=11, bold=True)
    
    ws['A10'] = "Posibles causas:"
    ws['B10'] = "1. No hay asignaciones en esta subdependencia"
    ws['B11'] = "2. Los bienes están marcados como 'devuelto'"
    ws['B12'] = "3. No hay detalles de asignación asociados"
    
    # Ajustar anchos
    for col in range(1, 8):
        ws.column_dimensions[get_column_letter(col)].width = 20
    
    wb.save(response)
    return response
def configurar_hoja_resumen(ws, subdependencia, datos_por_responsable, estilos):
    """Configura la hoja de resumen/índice"""
    
    # Configurar anchos
    anchos = {'A': 8, 'B': 40, 'C': 20, 'D': 25}
    for col, ancho in anchos.items():
        ws.column_dimensions[col].width = ancho
    
    row_num = 1
    
    # Título principal
    fecha_actual = datetime.now().strftime("%d/%m/%Y")
    ws.merge_cells(f'A{row_num}:D{row_num}')
    ws.cell(row=row_num, column=1, 
           value=f"REPORTE DE SUBDEPENDENCIA - {subdependencia.nombre.upper()}")
    aplicar_estilo_celda(ws.cell(row=row_num, column=1), estilos['titulo_principal'])
    ws.row_dimensions[row_num].height = 35
    row_num += 1
    
    # Fecha
    ws.merge_cells(f'A{row_num}:D{row_num}')
    ws.cell(row=row_num, column=1, value=f"Fecha de generación: {fecha_actual}")
    ws.cell(row=row_num, column=1).font = Font(name='Calibri', size=10, italic=True)
    ws.row_dimensions[row_num].height = 20
    row_num += 2
    
    # Información general
    ws.cell(row=row_num, column=1, value="DEPENDENCIA:")
    aplicar_estilo_celda(ws.cell(row=row_num, column=1), estilos['etiqueta'])
    ws.merge_cells(f'B{row_num}:D{row_num}')
    ws.cell(row=row_num, column=2, value=subdependencia.dependencia.nombre)
    aplicar_estilo_celda(ws.cell(row=row_num, column=2), estilos['info_valor'])
    row_num += 1
    
    ws.cell(row=row_num, column=1, value="SUBDEPENDENCIA:")
    aplicar_estilo_celda(ws.cell(row=row_num, column=1), estilos['etiqueta'])
    ws.merge_cells(f'B{row_num}:D{row_num}')
    ws.cell(row=row_num, column=2, value=subdependencia.nombre)
    aplicar_estilo_celda(ws.cell(row=row_num, column=2), estilos['info_valor'])
    row_num += 2
    
    # Estadísticas rápidas
    total_responsables = len(datos_por_responsable)
    total_bienes = sum(len(data['bienes']) for data in datos_por_responsable.values())
    
    ws.cell(row=row_num, column=1, value="TOTAL RESPONSABLES:")
    aplicar_estilo_celda(ws.cell(row=row_num, column=1), estilos['etiqueta'])
    ws.cell(row=row_num, column=2, value=total_responsables)
    aplicar_estilo_celda(ws.cell(row=row_num, column=2), estilos['info_valor'])
    row_num += 1
    
    ws.cell(row=row_num, column=1, value="TOTAL BIENES:")
    aplicar_estilo_celda(ws.cell(row=row_num, column=1), estilos['etiqueta'])
    ws.cell(row=row_num, column=2, value=total_bienes)
    aplicar_estilo_celda(ws.cell(row=row_num, column=2), estilos['info_valor'])
    row_num += 3
    
    # Encabezados de la tabla de índice
    encabezados = ["No.", "USUARIO", "TOTAL BIENES", "ENLACE"]
    for col_num, encabezado in enumerate(encabezados, 1):
        celda = ws.cell(row=row_num, column=col_num, value=encabezado)
        aplicar_estilo_celda(celda, estilos['encabezado_tabla'])
    
    ws.row_dimensions[row_num].height = 25
    row_num += 1
    
    # Listar responsables con hipervínculos
    # Ordenar por nombre de responsable
    responsables_items = list(datos_por_responsable.items())
    responsables_items.sort(key=lambda x: x[1]['nombre'])
    
    for idx, (responsable, data) in enumerate(responsables_items, 1):
        nombre_responsable = data['nombre']
        bienes = data['bienes']
        nombre_hoja = sanitizar_nombre_hoja(nombre_responsable)
        
        # --- FRAGMENTO ESPECÍFICO PARA EL ENLACE ---
        # Buscar la hoja real (puede tener sufijo por nombres duplicados)
        nombre_hoja_real = None
        for sheet_name in ws.parent.sheetnames:
            # Verificar si la hoja existe (comparar sin espacios/guiones)
            nombre_base = nombre_responsable.replace(' ', '_').lower()
            sheet_base = sheet_name.replace(' ', '_').lower()
            
            if nombre_base in sheet_base or sheet_name.startswith(nombre_hoja.split('_')[0]):
                nombre_hoja_real = sheet_name
                break
        
        # Estilo alternado para filas
        if idx % 2 == 0:
            estilo_fila = estilos['celda_datos_alterna']
        else:
            estilo_fila = estilos['celda_datos']
        
        # Número
        celda_num = ws.cell(row=row_num, column=1, value=idx)
        aplicar_estilo_celda(celda_num, estilo_fila)
        
        # --- NOMBRE CON HIPERVÍNCULO ---
        celda_nombre = ws.cell(row=row_num, column=2, value=nombre_responsable)
        
        # Aplicar estilo base
        celda_nombre.font = Font(name='Calibri', size=10, color="000000")
        celda_nombre.alignment = Alignment(horizontal="left", vertical="center")
        celda_nombre.border = estilo_fila['border']
        celda_nombre.fill = estilo_fila['fill']
        
        # Agregar hipervínculo si la hoja existe
        if nombre_hoja_real and nombre_hoja_real in ws.parent.sheetnames:
            celda_nombre.hyperlink = f"#{nombre_hoja_real}!A1"
            celda_nombre.font = Font(name='Calibri', size=10, color="2563EB", underline="single")
        # --- FIN HIPERVÍNCULO NOMBRE ---
        
        # Total de bienes
        celda_total = ws.cell(row=row_num, column=3, value=len(bienes))
        aplicar_estilo_celda(celda_total, estilo_fila)
        
        # --- ENLACE EN COLUMNA SEPARADA ---
        celda_enlace = ws.cell(row=row_num, column=4)
        
        if nombre_hoja_real and nombre_hoja_real in ws.parent.sheetnames:
            celda_enlace.value = "Ver detalle"
            celda_enlace.hyperlink = f"#{nombre_hoja_real}!A1"
            celda_enlace.font = Font(name='Calibri', size=10, color="2563EB", underline="single")
        else:
            celda_enlace.value = "-"
        
        # Aplicar estilo al enlace
        celda_enlace.alignment = Alignment(horizontal="center", vertical="center")
        celda_enlace.border = estilo_fila['border']
        celda_enlace.fill = estilo_fila['fill']
        # --- FIN ENLACE COLUMNA ---
        
        row_num += 1

def configurar_hoja_responsable(ws, nombre_responsable, bienes_responsable, subdependencia, estilos):
    """Configura una hoja para un responsable específico"""
    
    # Configurar anchos de columnas
    anchos = {
        'A': 20,   # Código
        'B': 25,   # Tipo
        'C': 20,   # Marca
        'D': 20,   # Modelo
        'E': 25,   # Serial
        'F': 20,   # Condición
    }
    
    for col, ancho in anchos.items():
        ws.column_dimensions[col].width = ancho
    
    row_num = 1
    
    # Título principal de la hoja
    fecha_actual = datetime.now().strftime("%d/%m/%Y")
    
    ws.merge_cells(f'A{row_num}:F{row_num}')
    ws.cell(row=row_num, column=1, 
           value=f"BIENES ASIGNADOS A: {nombre_responsable.upper()}")
    aplicar_estilo_celda(ws.cell(row=row_num, column=1), estilos['titulo_responsable'])
    ws.row_dimensions[row_num].height = 35
    row_num += 1
    
    # Fecha
    ws.merge_cells(f'A{row_num}:F{row_num}')
    ws.cell(row=row_num, column=1, value=f"Fecha: {fecha_actual}")
    ws.cell(row=row_num, column=1).font = Font(name='Calibri', size=10, italic=True)
    ws.row_dimensions[row_num].height = 20
    row_num += 2
    
    # Información del responsable
    ws.cell(row=row_num, column=1, value="RESPONSABLE:")
    aplicar_estilo_celda(ws.cell(row=row_num, column=1), estilos['etiqueta'])
    ws.merge_cells(f'B{row_num}:F{row_num}')
    ws.cell(row=row_num, column=2, value=nombre_responsable)
    aplicar_estilo_celda(ws.cell(row=row_num, column=2), estilos['info_valor'])
    row_num += 1
    
    ws.cell(row=row_num, column=1, value="DEPENDENCIA:")
    aplicar_estilo_celda(ws.cell(row=row_num, column=1), estilos['etiqueta'])
    ws.merge_cells(f'B{row_num}:F{row_num}')
    ws.cell(row=row_num, column=2, value=subdependencia.dependencia.nombre)
    aplicar_estilo_celda(ws.cell(row=row_num, column=2), estilos['info_valor'])
    row_num += 1
    
    ws.cell(row=row_num, column=1, value="SUBDEPENDENCIA:")
    aplicar_estilo_celda(ws.cell(row=row_num, column=1), estilos['etiqueta'])
    ws.merge_cells(f'B{row_num}:F{row_num}')
    ws.cell(row=row_num, column=2, value=subdependencia.nombre)
    aplicar_estilo_celda(ws.cell(row=row_num, column=2), estilos['info_valor'])
    row_num += 2
    
    # Encabezados de la tabla de bienes
    cabeceras = ['N° de inventario del bien', 'TIPO', 'MARCA', 'MODELO', 'SERIAL', 'CONDICIÓN']
    
    for col_num, cabecera in enumerate(cabeceras, 1):
        celda = ws.cell(row=row_num, column=col_num, value=cabecera)
        aplicar_estilo_celda(celda, estilos['encabezado_tabla'])
    
    ws.row_dimensions[row_num].height = 25
    row_num += 1
    
    # Escribir los bienes del responsable
    for idx, detalle in enumerate(bienes_responsable):
        # Determinar estilo de fila alternado
        if idx % 2 == 0:
            estilo_base = estilos['celda_datos_alterna']
            estilo_izquierda = estilos['celda_izquierda_alterna']
        else:
            estilo_base = estilos['celda_datos']
            estilo_izquierda = estilos['celda_izquierda']
        
        # Obtener datos del bien
        bien = detalle.bien
        
        # Preparar datos de la fila con validación
        datos_fila = [
            bien.codigo_bien if bien and bien.codigo_bien else "Sin código",
            bien.tipo_bien.nombre if bien and bien.tipo_bien and bien.tipo_bien.nombre else "Sin tipo",
            bien.marca.nombre if bien and bien.marca and bien.marca.nombre else "Sin marca",
            bien.modelo.nombre if bien and bien.modelo and bien.modelo.nombre else "Sin modelo",
            bien.serial if bien and bien.serial else "Sin serial",
            bien.condicion.nombre if bien and bien.condicion and bien.condicion.nombre else "Sin condición"
        ]
        
        # Escribir cada celda
        for col_num, valor in enumerate(datos_fila, 1):
            celda = ws.cell(row=row_num, column=col_num, value=str(valor))
            
            if col_num == 1:  # Código (centrado)
                aplicar_estilo_celda(celda, estilo_base)
            else:  # Resto de columnas (izquierda)
                aplicar_estilo_celda(celda, estilo_izquierda)
        
        row_num += 1
    
    # Fila de total
    ws.merge_cells(f'A{row_num}:E{row_num}')
    celda_total = ws.cell(row=row_num, column=1, value=f"TOTAL DE BIENES: {len(bienes_responsable)}")
    aplicar_estilo_celda(celda_total, estilos['total'])
    
    celda_valor = ws.cell(row=row_num, column=6, value=len(bienes_responsable))
    aplicar_estilo_celda(celda_valor, estilos['total'])
    
    row_num += 2
    
    # Información adicional
    ws.cell(row=row_num, column=1, value="Fecha de generación:")
    ws.cell(row=row_num, column=1).font = Font(name='Calibri', size=10, bold=True)
    
    ws.cell(row=row_num, column=2, value=datetime.now().strftime("%d/%m/%Y %H:%M:%S"))
    row_num += 1

def configurar_hoja_estadisticas(ws, subdependencia, datos_por_responsable, estilos):
    """Configura la hoja de estadísticas"""
    
    # Configurar anchos
    anchos = {'A': 30, 'B': 20}
    for col, ancho in anchos.items():
        ws.column_dimensions[col].width = ancho
    
    row_num = 1
    
    # Título
    ws.merge_cells(f'A{row_num}:B{row_num}')
    ws.cell(row=row_num, column=1, 
           value=f"ESTADÍSTICAS - {subdependencia.nombre.upper()}")
    aplicar_estilo_celda(ws.cell(row=row_num, column=1), estilos['titulo_principal'])
    ws.row_dimensions[row_num].height = 35
    row_num += 2
    
    # Información general
    ws.cell(row=row_num, column=1, value="DEPENDENCIA:")
    aplicar_estilo_celda(ws.cell(row=row_num, column=1), estilos['etiqueta'])
    ws.cell(row=row_num, column=2, value=subdependencia.dependencia.nombre)
    aplicar_estilo_celda(ws.cell(row=row_num, column=2), estilos['info_valor'])
    row_num += 1
    
    ws.cell(row=row_num, column=1, value="SUBDEPENDENCIA:")
    aplicar_estilo_celda(ws.cell(row=row_num, column=1), estilos['etiqueta'])
    ws.cell(row=row_num, column=2, value=subdependencia.nombre)
    aplicar_estilo_celda(ws.cell(row=row_num, column=2), estilos['info_valor'])
    row_num += 2
    
    # Estadísticas
    total_responsables = len(datos_por_responsable)
    total_bienes = sum(len(data['bienes']) for data in datos_por_responsable.values())
    promedio_bienes = total_bienes / total_responsables if total_responsables > 0 else 0
    
    ws.cell(row=row_num, column=1, value="Total de Responsables:")
    aplicar_estilo_celda(ws.cell(row=row_num, column=1), estilos['etiqueta'])
    ws.cell(row=row_num, column=2, value=total_responsables)
    aplicar_estilo_celda(ws.cell(row=row_num, column=2), estilos['info_valor'])
    row_num += 1
    
    ws.cell(row=row_num, column=1, value="Total de Bienes:")
    aplicar_estilo_celda(ws.cell(row=row_num, column=1), estilos['etiqueta'])
    ws.cell(row=row_num, column=2, value=total_bienes)
    aplicar_estilo_celda(ws.cell(row=row_num, column=2), estilos['info_valor'])
    row_num += 1
    
    ws.cell(row=row_num, column=1, value="Promedio de bienes por responsable:")
    aplicar_estilo_celda(ws.cell(row=row_num, column=1), estilos['etiqueta'])
    ws.cell(row=row_num, column=2, value=round(promedio_bienes, 2))
    aplicar_estilo_celda(ws.cell(row=row_num, column=2), estilos['info_valor'])
    row_num += 3
    
    # Distribución de bienes por responsable
    ws.cell(row=row_num, column=1, value="Distribución de Bienes:")
    ws.cell(row=row_num, column=1).font = Font(name='Calibri', size=11, bold=True)
    row_num += 1
    
    # Encabezados
    ws.cell(row=row_num, column=1, value="Responsable")
    aplicar_estilo_celda(ws.cell(row=row_num, column=1), estilos['encabezado_tabla'])
    ws.cell(row=row_num, column=2, value="Cantidad")
    aplicar_estilo_celda(ws.cell(row=row_num, column=2), estilos['encabezado_tabla'])
    row_num += 1
    
    # Listar responsables ordenados por cantidad de bienes
    responsables_items = list(datos_por_responsable.items())
    responsables_items.sort(key=lambda x: len(x[1]['bienes']), reverse=True)
    
    for idx, (responsable, data) in enumerate(responsables_items, 1):
        nombre_responsable = data['nombre']
        
        # Estilo alternado
        if idx % 2 == 0:
            estilo_fila = estilos['celda_datos_alterna']
        else:
            estilo_fila = estilos['celda_datos']
        
        # Responsable
        celda_responsable = ws.cell(row=row_num, column=1, value=nombre_responsable)
        aplicar_estilo_celda(celda_responsable, estilo_fila)
        
        # Cantidad
        celda_cantidad = ws.cell(row=row_num, column=2, value=len(data['bienes']))
        aplicar_estilo_celda(celda_cantidad, estilo_fila)
        
        row_num += 1


def exportar_bienes_por_subdependencia_excel(request, subdependencia_id=None):
    """
    Genera un reporte en Excel específico para una subdependencia.
    Una pestaña por cada responsable con sus bienes a cargo.
    """
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    fecha_descarga = datetime.now().strftime("%Y%m%d_%H%M")
    response['Content-Disposition'] = f'attachment; filename="reporte_subdependencia_{fecha_descarga}.xlsx"'
    
    print(f"=== INICIANDO REPORTE PARA SUBDEPENDENCIA ID: {subdependencia_id} ===")
    
    # Validar si se proporcionó un ID de subdependencia
    if subdependencia_id is None:
        print("ERROR: No se proporcionó ID de subdependencia")
        return crear_reporte_error(response, "Debe especificar una subdependencia")
    
    # Obtener la subdependencia
    try:
        subdependencia = Subdependencia.objects.get(id=subdependencia_id)
        print(f"✓ Subdependencia encontrada: {subdependencia.nombre} (ID: {subdependencia.id})")
        print(f"✓ Dependencia padre: {subdependencia.dependencia.nombre}")
    except Subdependencia.DoesNotExist:
        print(f"✗ Subdependencia con ID {subdependencia_id} no encontrada")
        return crear_reporte_error(response, f"Subdependencia con ID {subdependencia_id} no encontrada")
    
    # DEBUG: Verificar asignaciones en esta subdependencia
    print(f"\n=== VERIFICANDO ASIGNACIONES EN LA SUBDEPENDENCIA ===")
    asignaciones_count = Asignacion.objects.filter(subdependencia=subdependencia).count()
    print(f"Total de asignaciones en la subdependencia: {asignaciones_count}")
    
    # Listar asignaciones para debug
    asignaciones = Asignacion.objects.filter(subdependencia=subdependencia).select_related('usuario')
    for i, asignacion in enumerate(asignaciones[:5], 1):  # Mostrar primeras 5
        usuario = asignacion.usuario.nombres_apellidos if asignacion.usuario else "Sin usuario"
        print(f"  {i}. Asignación {asignacion.nro_asignacion} - Usuario: {usuario}")
    
    # DEBUG: Verificar detalles de asignación sin filtros primero
    print(f"\n=== VERIFICANDO DETALLES DE ASIGNACIÓN ===")
    detalles_count = DetalleAsignacion.objects.filter(asignacion__subdependencia=subdependencia).count()
    print(f"Total de detalles de asignación (incluyendo devueltos): {detalles_count}")
    
    # DEBUG: Contar detalles no devueltos
    detalles_no_devueltos = DetalleAsignacion.objects.filter(
        asignacion__subdependencia=subdependencia,
        devuelto=False
    ).count()
    print(f"Total de detalles NO devueltos: {detalles_no_devueltos}")
    
    # Obtener datos de bienes asignados en esta subdependencia - VERSIÓN CORREGIDA
    print(f"\n=== OBTENIENDO DATOS PARA EL REPORTE ===")
    
    # PRIMERO: Intentar con la consulta original
    detalles_asignacion = DetalleAsignacion.objects.filter(
        asignacion__subdependencia=subdependencia,
        devuelto=False
    ).select_related(
        'bien',
        'bien__tipo_bien',
        'bien__marca',
        'bien__modelo',
        'bien__condicion',
        'asignacion',
        'asignacion__dependencia',
        'asignacion__subdependencia',
        'asignacion__usuario'
    ).order_by('asignacion__usuario__nombres_apellidos')
    
    conteo = detalles_asignacion.count()
    print(f"Consulta 1 (con filtro devuelto=False): {conteo} registros")
    
    # SEGUNDO: Si no hay datos, intentar sin filtrar por devuelto
    if conteo == 0:
        print("✓ Intentando sin filtrar por devuelto=False...")
        detalles_asignacion = DetalleAsignacion.objects.filter(
            asignacion__subdependencia=subdependencia
        ).select_related(
            'bien',
            'bien__tipo_bien',
            'bien__marca',
            'bien__modelo',
            'bien__condicion',
            'asignacion',
            'asignacion__dependencia',
            'asignacion__subdependencia',
            'asignacion__usuario'
        ).order_by('asignacion__usuario__nombres_apellidos')
        
        conteo = detalles_asignacion.count()
        print(f"Consulta 2 (sin filtro devuelto): {conteo} registros")
    
    # TERCERO: Si aún no hay datos, probar una consulta más básica
    if conteo == 0:
        print("✓ Intentando consulta básica...")
        detalles_asignacion = DetalleAsignacion.objects.filter(
            asignacion__subdependencia=subdependencia
        ).select_related('bien', 'asignacion', 'asignacion__usuario')
        
        conteo = detalles_asignacion.count()
        print(f"Consulta 3 (básica): {conteo} registros")
        
        # Mostrar algunos detalles para debug
        if conteo > 0:
            print("\n✓ Mostrando primeros 3 detalles encontrados:")
            for i, detalle in enumerate(detalles_asignacion[:3], 1):
                print(f"  {i}. ID: {detalle.id}")
                print(f"     Bien: {detalle.bien.codigo_bien if detalle.bien else 'Sin bien'}")
                print(f"     Devuelto: {detalle.devuelto}")
                if detalle.asignacion:
                    print(f"     Asignación: {detalle.asignacion.nro_asignacion}")
                    print(f"     Usuario: {detalle.asignacion.usuario.nombres_apellidos if detalle.asignacion.usuario else 'Sin usuario'}")
                print()
    
    # Si aún no hay datos, verificar si hay problemas con las relaciones
    if conteo == 0:
        print("\n⚠️  NO SE ENCONTRARON DETALLES DE ASIGNACIÓN")
        print("Posibles causas:")
        print("1. No hay asignaciones en esta subdependencia")
        print("2. Todos los bienes están marcados como 'devuelto=True'")
        print("3. Problemas con las relaciones entre modelos")
        
        # Verificar si hay asignaciones pero sin detalles
        asignaciones_con_detalles = Asignacion.objects.filter(
            subdependencia=subdependencia,
            detalleasignacion__isnull=False
        ).distinct().count()
        print(f"Asignaciones con detalles: {asignaciones_con_detalles}")
    
    # Si no hay datos, crear reporte informativo
    if conteo == 0:
        print("\n✓ Creando reporte informativo (sin datos)...")
        return crear_reporte_subdependencia_vacia(response, subdependencia)
    
    print(f"\n✓ Procesando {conteo} detalles de asignación...")
    
    # Crear el libro de Excel
    wb = openpyxl.Workbook()
    
    # Eliminar la hoja por defecto
    if 'Sheet' in wb.sheetnames:
        ws_default = wb['Sheet']
        wb.remove(ws_default)
    
    estilos = get_estilos()
    
    # Agrupar datos por responsable
    print("✓ Agrupando datos por responsable...")
    datos_por_responsable = {}
    
    for detalle in detalles_asignacion:
        if detalle.asignacion and detalle.asignacion.usuario:
            responsable = detalle.asignacion.usuario
            responsable_nombre = responsable.nombres_apellidos if responsable.nombres_apellidos else f"Responsable_{responsable.id}"
        else:
            responsable_nombre = "Sin_Responsable"
            responsable = None
        
        if responsable not in datos_por_responsable:
            datos_por_responsable[responsable] = {
                'nombre': responsable_nombre,
                'bienes': []
            }
        
        datos_por_responsable[responsable]['bienes'].append(detalle)
    
    print(f"✓ Responsables encontrados: {len(datos_por_responsable)}")
    
    # Mostrar información de cada responsable para debug
    for responsable, data in datos_por_responsable.items():
        nombre = data['nombre']
        cantidad = len(data['bienes'])
        print(f"  - {nombre}: {cantidad} bienes")
    

    
    # Crear una hoja por cada responsable
    for responsable, data in datos_por_responsable.items():
        nombre_responsable = data['nombre']
        bienes_responsable = data['bienes']
        
        # Crear nombre válido para la hoja
        nombre_hoja = sanitizar_nombre_hoja(nombre_responsable)
        
        # Si ya existe una hoja con ese nombre, agregar sufijo
        nombre_original = nombre_hoja
        contador = 1
        while nombre_hoja in wb.sheetnames:
            nombre_hoja = f"{nombre_original}_{contador}"
            contador += 1
        
        print(f"✓ Creando hoja para: {nombre_responsable} -> {nombre_hoja}")
        
        # Crear nueva hoja
        ws = wb.create_sheet(title=nombre_hoja)
        configurar_hoja_responsable(ws, nombre_responsable, bienes_responsable, subdependencia, estilos)

            # Crear hoja de resumen/índice primero
    ws_resumen = wb.create_sheet(title="Índice")
    configurar_hoja_resumen(ws_resumen, subdependencia, datos_por_responsable, estilos)
    # # Crear hoja de estadísticas
    # ws_estadisticas = wb.create_sheet(title="Estadísticas")
    # configurar_hoja_estadisticas(ws_estadisticas, subdependencia, datos_por_responsable, estilos)
    
    # Mover hojas en orden lógico
    reordenar_hojas(wb)
    
    # Configurar hoja activa
    wb.active = wb["Índice"]
    
    # Guardar
    wb.save(response)
    
    print(f"\n=== REPORTE GENERADO EXITOSAMENTE ===")
    print(f"Archivo: reporte_subdependencia_{fecha_descarga}.xlsx")
    print(f"Total responsables: {len(datos_por_responsable)}")
    print(f"Total bienes: {conteo}")
    
    return response

# --- VERSIÓN PARA USAR EN URLS ---

def reporte_subdependencia_view(request, subdependencia_id):
    """Wrapper para usar en URLs"""
    return exportar_bienes_por_subdependencia_excel(request, subdependencia_id)

def reporte_subdependencia_actual(request):
    """Versión para obtener subdependencia de los parámetros GET"""
    subdependencia_id = request.GET.get('subdependencia_id')
    if not subdependencia_id:
        # Intentar obtener del parámetro en la URL
        from django.urls import resolve
        try:
            match = resolve(request.path_info)
            subdependencia_id = match.kwargs.get('subdependencia_id')
        except:
            pass
    
    if not subdependencia_id:
        return crear_reporte_error(HttpResponse(), "Debe especificar una subdependencia")
    
    return exportar_bienes_por_subdependencia_excel(request, subdependencia_id)