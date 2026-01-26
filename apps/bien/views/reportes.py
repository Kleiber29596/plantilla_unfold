import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from apps.bien.models.detalle_asignacion import DetalleAsignacion
from apps.auxiliares.models.dependencia import Dependencia
from itertools import groupby
from datetime import datetime
import re
import logging

# Configurar logging para debug
logger = logging.getLogger(__name__)

# --- Funciones de Estilo CORREGIDAS ---

def get_estilos():
    """Retorna un diccionario con todos los estilos visuales mejorados"""
    
    COLOR_PRIMARIO = "B7B7B7"
    COLOR_SECUNDARIO = "B7B7B7"
    COLOR_FONDO_CLARO = "F8FAFC"
    COLOR_TEXTO_BLANCO = "FFFFFF"
    COLOR_TEXTO_OSCURO = "1F2937"
    COLOR_BORDE = "CBD5E1"
    COLOR_ENCABEZADO = "B7B7B7"
    COLOR_ENCABEZADO_SECCION = "D9D9D9"
    
    borde_delgado = Border(
        left=Side(style='thin', color=COLOR_BORDE),
        right=Side(style='thin', color=COLOR_BORDE),
        top=Side(style='thin', color=COLOR_BORDE),
        bottom=Side(style='thin', color=COLOR_BORDE)
    )
    
    borde_grueso = Border(
        left=Side(style='medium', color=COLOR_PRIMARIO),
        right=Side(style='medium', color=COLOR_PRIMARIO),
        top=Side(style='medium', color=COLOR_PRIMARIO),
        bottom=Side(style='medium', color=COLOR_PRIMARIO)
    )
    
    # Crear objetos PatternFill explícitamente
    fill_primario = PatternFill(start_color=COLOR_PRIMARIO, end_color=COLOR_PRIMARIO, fill_type="solid")
    fill_encabezado = PatternFill(start_color=COLOR_ENCABEZADO, end_color=COLOR_ENCABEZADO, fill_type="solid")
    fill_fondo_claro = PatternFill(start_color=COLOR_FONDO_CLARO, end_color=COLOR_FONDO_CLARO, fill_type="solid")
    fill_encabezado_seccion = PatternFill(start_color=COLOR_ENCABEZADO_SECCION, end_color=COLOR_ENCABEZADO_SECCION, fill_type="solid")
    
    return {
        'titulo_principal': {
            'font': Font(name='Calibri', size=16, bold=True, color=COLOR_TEXTO_OSCURO),
            'fill': fill_primario,
            'alignment': Alignment(horizontal="center", vertical="center", wrap_text=True),
            'border': borde_grueso
        },
        'encabezado_tabla': {
            'font': Font(name='Calibri', size=11, bold=True, color=COLOR_TEXTO_OSCURO),
            'fill': fill_encabezado,
            'alignment': Alignment(horizontal="center", vertical="center", wrap_text=True),
            'border': borde_delgado
        },
        'celda_datos': {
            'font': Font(name='Calibri', size=10, color=COLOR_TEXTO_OSCURO),
            'fill': PatternFill(fill_type=None),  # Fill vacío explícito
            'alignment': Alignment(horizontal="center", vertical="center"),
            'border': borde_delgado
        },
        'celda_datos_alterna': {
            'font': Font(name='Calibri', size=10, color=COLOR_TEXTO_OSCURO),
            'fill': fill_fondo_claro,
            'alignment': Alignment(horizontal="center", vertical="center"),
            'border': borde_delgado
        },
        'etiqueta': {
            'font': Font(name='Calibri', size=11, bold=True, color=COLOR_TEXTO_OSCURO),
            'fill': PatternFill(fill_type=None),
            'alignment': Alignment(horizontal="left", vertical="center"),
            'border': borde_delgado
        },
        'info_valor': {
            'font': Font(name='Calibri', size=11, color=COLOR_TEXTO_OSCURO),
            'fill': PatternFill(fill_type=None),
            'alignment': Alignment(horizontal="left", vertical="center"),
            'border': borde_delgado
        },
        'dato_izquierda': {
            'font': Font(name='Calibri', size=10, color=COLOR_TEXTO_OSCURO),
            'fill': PatternFill(fill_type=None),
            'alignment': Alignment(horizontal="left", vertical="center", wrap_text=True),
            'border': borde_delgado
        },
        'dato_izquierda_alterna': {
            'font': Font(name='Calibri', size=10, color=COLOR_TEXTO_OSCURO),
            'fill': fill_fondo_claro,
            'alignment': Alignment(horizontal="left", vertical="center", wrap_text=True),
            'border': borde_delgado
        },
    }

def aplicar_estilo_celda(celda, estilo):
    """Aplica un estilo completo a una celda de manera segura"""
    if estilo:
        if 'font' in estilo and estilo['font']:
            celda.font = estilo['font']
        if 'fill' in estilo and estilo['fill']:
            celda.fill = estilo['fill']
        if 'alignment' in estilo and estilo['alignment']:
            celda.alignment = estilo['alignment']
        if 'border' in estilo and estilo['border']:
            celda.border = estilo['border']

def sanitizar_nombre_hoja(nombre):
    """Limpia el nombre para que sea válido como nombre de hoja en Excel"""
    if not nombre:
        return "SinNombre"
    
    nombre = str(nombre)
    nombre = re.sub(r'[\\/*?:[\]]', '_', nombre)
    if len(nombre) > 31:
        nombre = nombre[:28] + "..."
    return nombre

# --- FUNCIÓN PRINCIPAL CORREGIDA ---

def exportar_bienes_por_dependencia_excel(request, dependencia_id=None):
    """
    Genera un reporte en Excel de bienes asignados para una dependencia específica.
    Una pestaña por cada subdependencia, con responsables como columna.
    """
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    fecha_descarga = datetime.now().strftime("%Y%m%d_%H%M")
    response['Content-Disposition'] = f'attachment; filename="reporte_inventario_{fecha_descarga}.xlsx"'
    
    logger.info(f"Iniciando generación de reporte - dependencia_id: {dependencia_id}")
    
    # Obtener la dependencia
    dependencia = None
    if dependencia_id:
        try:
            dependencia = Dependencia.objects.get(id=dependencia_id)
            logger.info(f"Dependencia encontrada por ID {dependencia_id}: {dependencia.nombre}")
        except Dependencia.DoesNotExist:
            logger.warning(f"Dependencia con ID {dependencia_id} no encontrada")
            dependencia = Dependencia.objects.first()
    else:
        dependencia = Dependencia.objects.first()
    
    if not dependencia:
        logger.error("No hay dependencias configuradas en el sistema")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sin Datos"
        ws['A1'] = "ERROR: No hay dependencias configuradas en el sistema"
        wb.save(response)
        return response
    
    logger.info(f"Usando dependencia: {dependencia.nombre} (ID: {dependencia.id})")
    
    # Obtener todas las subdependencias de esta dependencia
    subdependencias = dependencia.subdependencia_set.all().order_by('nombre')
    logger.info(f"Subdependencias encontradas: {subdependencias.count()}")
    
    if not subdependencias.exists():
        logger.warning(f"No hay subdependencias para {dependencia.nombre}")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sin Datos"
        ws['A1'] = f"No hay subdependencias para {dependencia.nombre}"
        wb.save(response)
        return response
    
    # Obtener datos de bienes asignados
    try:
        logger.info("Obteniendo datos de bienes asignados...")
        
        # OPCIÓN MÁS SEGURA: Obtener todos los datos sin filtros complejos primero
        detalles_asignacion = DetalleAsignacion.objects.select_related(
            'bien',
            'bien__tipo_bien',
            'bien__marca',
            'bien__modelo',
            'bien__condicion',
            'asignacion',
            'asignacion__dependencia',
            'asignacion__subdependencia',
            'asignacion__usuario'
        ).all()
        
        # Filtrar por dependencia
        detalles_asignacion = [d for d in detalles_asignacion 
                              if d.asignacion and d.asignacion.dependencia == dependencia]
        
        conteo = len(detalles_asignacion)
        logger.info(f"Detalles encontrados en la dependencia: {conteo}")
        
    except Exception as e:
        logger.error(f"Error al obtener datos: {str(e)}")
        detalles_asignacion = []
        conteo = 0
    
    # Si no hay datos, crear reporte informativo
    if not detalles_asignacion:
        logger.warning("No hay datos de detalles de asignación")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Información"
        
        estilos = get_estilos()
        
        ws['A1'] = f"REPORTE DE INVENTARIO - {dependencia.nombre.upper()}"
        ws.merge_cells('A1:G1')
        aplicar_estilo_celda(ws['A1'], estilos['titulo_principal'])
        ws.row_dimensions[1].height = 35
        
        ws['A3'] = "INFORMACIÓN DEL SISTEMA"
        ws.merge_cells('A3:G3')
        aplicar_estilo_celda(ws['A3'], estilos['encabezado_tabla'])
        
        ws['A5'] = "Dependencia seleccionada:"
        aplicar_estilo_celda(ws['A5'], estilos['etiqueta'])
        ws['B5'] = dependencia.nombre
        aplicar_estilo_celda(ws['B5'], estilos['info_valor'])
        
        ws['A6'] = "Subdependencias encontradas:"
        aplicar_estilo_celda(ws['A6'], estilos['etiqueta'])
        ws['B6'] = subdependencias.count()
        aplicar_estilo_celda(ws['B6'], estilos['info_valor'])
        
        ws['A8'] = "No se encontraron bienes asignados en esta dependencia."
        ws.merge_cells('A8:G8')
        celda_mensaje = ws['A8']
        celda_mensaje.font = Font(name='Calibri', size=12, bold=True, color="FF0000")
        celda_mensaje.alignment = Alignment(horizontal="center", vertical="center")
        
        # Ajustar anchos
        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 20
        
        wb.save(response)
        return response
    
    logger.info(f"Procesando {conteo} detalles de asignación...")
    
    # Crear el libro de Excel
    wb = openpyxl.Workbook()
    
    # Eliminar la hoja por defecto
    if 'Sheet' in wb.sheetnames:
        ws_default = wb['Sheet']
        wb.remove(ws_default)
    
    estilos = get_estilos()
    
    # Agrupar datos por subdependencia
    datos_por_subdependencia = {}
    for detalle in detalles_asignacion:
        if detalle.asignacion and detalle.asignacion.subdependencia:
            subdependencia = detalle.asignacion.subdependencia
            if subdependencia not in datos_por_subdependencia:
                datos_por_subdependencia[subdependencia] = []
            datos_por_subdependencia[subdependencia].append(detalle)
    
    logger.info(f"Datos agrupados por {len(datos_por_subdependencia)} subdependencias")
    
    # Crear una hoja por cada subdependencia
    for subdependencia in subdependencias:
        nombre_hoja = sanitizar_nombre_hoja(subdependencia.nombre)
        ws = wb.create_sheet(title=nombre_hoja)
        
        # Configurar anchos de columnas
        anchos = {
            'A': 20,   # Código
            'B': 25,   # Tipo
            'C': 20,   # Marca
            'D': 20,   # Modelo
            'E': 25,   # Serial
            'F': 20,   # Condición
            'G': 30,   # Responsable
        }
        
        for col, ancho in anchos.items():
            ws.column_dimensions[col].width = ancho
        
        row_num = 1
        
        # Título principal
        fecha_actual = datetime.now().strftime("%d/%m/%Y")
        ws.merge_cells(f'A{row_num}:G{row_num}')
        celda_titulo = ws.cell(row=row_num, column=1, 
                              value=f"INVENTARIO DE BIENES - {subdependencia.nombre.upper()} - {fecha_actual}")
        aplicar_estilo_celda(celda_titulo, estilos['titulo_principal'])
        ws.row_dimensions[row_num].height = 35
        row_num += 2
        
        # Información de dependencia
        ws.cell(row=row_num, column=1, value="DEPENDENCIA:")
        aplicar_estilo_celda(ws.cell(row=row_num, column=1), estilos['etiqueta'])
        ws.merge_cells(f'B{row_num}:G{row_num}')
        ws.cell(row=row_num, column=2, value=dependencia.nombre)
        aplicar_estilo_celda(ws.cell(row=row_num, column=2), estilos['info_valor'])
        row_num += 1
        
        # Información de subdependencia
        ws.cell(row=row_num, column=1, value="SUBDEPENDENCIA:")
        aplicar_estilo_celda(ws.cell(row=row_num, column=1), estilos['etiqueta'])
        ws.merge_cells(f'B{row_num}:G{row_num}')
        ws.cell(row=row_num, column=2, value=subdependencia.nombre)
        aplicar_estilo_celda(ws.cell(row=row_num, column=2), estilos['info_valor'])
        row_num += 2
        
        # Obtener bienes para esta subdependencia
        bienes_subdependencia = datos_por_subdependencia.get(subdependencia, [])
        
        if bienes_subdependencia:
            # Encabezados de tabla
            cabeceras = ['N° DE INVENTARIO DEL BIEN', 'TIPO', 'MARCA', 'MODELO', 'CONDICIÓN', 'USUARIO']
            
            for col_num, cabecera in enumerate(cabeceras, 1):
                celda = ws.cell(row=row_num, column=col_num, value=cabecera)
                aplicar_estilo_celda(celda, estilos['encabezado_tabla'])
            
            ws.row_dimensions[row_num].height = 25
            row_num += 1
            
            # Escribir datos
            for idx, detalle in enumerate(bienes_subdependencia):
                # Determinar estilos alternados
                if idx % 2 == 0:
                    estilo_base = estilos['celda_datos_alterna']
                    estilo_izquierda = estilos['dato_izquierda_alterna']
                else:
                    estilo_base = estilos['celda_datos']
                    estilo_izquierda = estilos['dato_izquierda']
                
                # Obtener datos del bien
                bien = detalle.bien
                asignacion = detalle.asignacion
                
                # Preparar datos de la fila
                datos_fila = [
                    bien.codigo_bien if bien and bien.codigo_bien else "Sin código",
                    bien.tipo_bien.nombre if bien and bien.tipo_bien and bien.tipo_bien.nombre else "Sin tipo",
                    bien.marca.nombre if bien and bien.marca and bien.marca.nombre else "Sin marca",
                    bien.modelo.nombre if bien and bien.modelo and bien.modelo.nombre else "Sin modelo",
                    bien.condicion.nombre if bien and bien.condicion and bien.condicion.nombre else "Sin condición",
                    asignacion.usuario.nombres_apellidos if asignacion and asignacion.usuario and asignacion.usuario.nombres_apellidos else "Sin asignar"
                ]
                
                # Escribir cada celda
                for col_num, valor in enumerate(datos_fila, 1):
                    celda = ws.cell(row=row_num, column=col_num, value=str(valor))
                    
                    if col_num == 1:  # Código (centrado)
                        aplicar_estilo_celda(celda, estilo_base)
                    else:  # Resto de columnas (izquierda)
                        aplicar_estilo_celda(celda, estilo_izquierda)
                
                row_num += 1
            
            # Total de bienes al final
            ws.merge_cells(f'A{row_num}:F{row_num}')
            total_celda = ws.cell(row=row_num, column=1, value=f"TOTAL DE BIENES: {len(bienes_subdependencia)}")
            aplicar_estilo_celda(total_celda, estilos['etiqueta'])
            row_num += 1
            
        else:
            # Mensaje si no hay bienes
            ws.merge_cells(f'A{row_num}:G{row_num}')
            celda_mensaje = ws.cell(row=row_num, column=1, value="NO HAY BIENES ASIGNADOS EN ESTA SUBDEPENDENCIA")
            aplicar_estilo_celda(celda_mensaje, estilos['encabezado_tabla'])
            ws.row_dimensions[row_num].height = 25
            row_num += 1
    
    # Crear hoja de resumen
    ws_resumen = wb.create_sheet(title="Resumen")
    
    # Configurar anchos
    anchos_resumen = {'A': 8, 'B': 40, 'C': 15, 'D': 25}
    for col, ancho in anchos_resumen.items():
        ws_resumen.column_dimensions[col].width = ancho
    
    row_num = 1
    
    # Título del resumen
    ws_resumen.merge_cells(f'A{row_num}:D{row_num}')
    titulo_resumen = ws_resumen.cell(row=row_num, column=1, 
                                   value=f"RESUMEN DE INVENTARIO - {dependencia.nombre.upper()}")
    aplicar_estilo_celda(titulo_resumen, estilos['titulo_principal'])
    ws_resumen.row_dimensions[row_num].height = 35
    row_num += 2
    
    # Encabezados del resumen
    encabezados = ["No.", "SUBDEPENDENCIA", "TOTAL BIENES", "ENLACE"]
    for col_num, encabezado in enumerate(encabezados, 1):
        celda = ws_resumen.cell(row=row_num, column=col_num, value=encabezado)
        aplicar_estilo_celda(celda, estilos['encabezado_tabla'])
    
    ws_resumen.row_dimensions[row_num].height = 25
    row_num += 1
    
    # Datos del resumen
    total_general = 0
    for idx, subdependencia in enumerate(subdependencias, 1):
        nombre_hoja = sanitizar_nombre_hoja(subdependencia.nombre)
        bienes_subdependencia = datos_por_subdependencia.get(subdependencia, [])
        total_bienes = len(bienes_subdependencia)
        total_general += total_bienes
        
        # Estilo alternado
        if idx % 2 == 0:
            estilo_fila = estilos['celda_datos_alterna']
        else:
            estilo_fila = estilos['celda_datos']
        
        # Número
        celda_num = ws_resumen.cell(row=row_num, column=1, value=idx)
        aplicar_estilo_celda(celda_num, estilo_fila)
        
        # Nombre de subdependencia
        celda_nombre = ws_resumen.cell(row=row_num, column=2, value=subdependencia.nombre)
        
        # Copiar estilo de la celda de datos pero mantener el texto negro
        celda_nombre.font = Font(name='Calibri', size=10, color="000000")
        celda_nombre.alignment = Alignment(horizontal="left", vertical="center")
        celda_nombre.border = estilo_fila['border']
        
        # Solo aplicar fill si existe en el estilo
        if 'fill' in estilo_fila and estilo_fila['fill']:
            celda_nombre.fill = estilo_fila['fill']
        
        # Agregar hipervínculo si tiene bienes
        if total_bienes > 0 and nombre_hoja in wb.sheetnames:
            celda_nombre.hyperlink = f"#{nombre_hoja}!A1"
            celda_nombre.font = Font(name='Calibri', size=10, color="2563EB", underline="single")
        
        # Total de bienes
        celda_total = ws_resumen.cell(row=row_num, column=3, value=total_bienes)
        aplicar_estilo_celda(celda_total, estilo_fila)
        
        # Enlace (si hay bienes)
        celda_enlace = ws_resumen.cell(row=row_num, column=4)
        
        if total_bienes > 0 and nombre_hoja in wb.sheetnames:
            celda_enlace.value = "Ver detalle"
            celda_enlace.hyperlink = f"#{nombre_hoja}!A1"
            celda_enlace.font = Font(name='Calibri', size=10, color="2563EB", underline="single")
        else:
            celda_enlace.value = "Sin bienes"
            celda_enlace.font = Font(name='Calibri', size=10, color="000000")
        
        # Aplicar estilo base
        celda_enlace.alignment = Alignment(horizontal="center", vertical="center")
        celda_enlace.border = estilo_fila['border']
        
        # Solo aplicar fill si existe
        if 'fill' in estilo_fila and estilo_fila['fill']:
            celda_enlace.fill = estilo_fila['fill']
        
        row_num += 1
    
    # Total general
    row_num += 1
    ws_resumen.merge_cells(f'A{row_num}:B{row_num}')
    celda_total_label = ws_resumen.cell(row=row_num, column=1, value="TOTAL GENERAL:")
    aplicar_estilo_celda(celda_total_label, estilos['etiqueta'])
    
    celda_total_valor = ws_resumen.cell(row=row_num, column=3, value=total_general)
    aplicar_estilo_celda(celda_total_valor, estilos['info_valor'])
    
    # Fecha de generación
    row_num += 2
    ws_resumen.cell(row=row_num, column=1, value="Fecha de generación:")
    aplicar_estilo_celda(ws_resumen.cell(row=row_num, column=1), estilos['etiqueta'])
    
    ws_resumen.cell(row=row_num, column=2, value=datetime.now().strftime("%d/%m/%Y %H:%M:%S"))
    aplicar_estilo_celda(ws_resumen.cell(row=row_num, column=2), estilos['info_valor'])
    
    # Mover resumen al principio
    resumen_sheet = wb["Resumen"]
    wb._sheets.insert(0, wb._sheets.pop(wb._sheets.index(resumen_sheet)))
    
    # Configurar hoja activa
    wb.active = wb.worksheets[0]
    
    # Guardar
    wb.save(response)
    
    logger.info(f"Reporte generado exitosamente. Total general: {total_general} bienes")
    return response


# --- VERSIÓN SIMPLIFICADA PARA DEPENDENCIA PREDETERMINADA ---

def exportar_bienes_dependencia_predeterminada(request):
    """
    Versión que busca la dependencia por nombre o usa la primera disponible
    """
    logger.info("Generando reporte para dependencia predeterminada")
    
    # Opción 1: Buscar por nombre - CAMBIA ESTO al nombre de tu dependencia
    nombre_dependencia_predeterminada = "NOMBRE_DE_TU_DEPENDENCIA_AQUI"  # ¡CAMBIA ESTO!
    
    try:
        dependencia = Dependencia.objects.get(nombre__icontains=nombre_dependencia_predeterminada)
        logger.info(f"Usando dependencia predeterminada por nombre: {dependencia.nombre}")
        return exportar_bienes_por_dependencia_excel(request, dependencia_id=dependencia.id)
    except Dependencia.DoesNotExist:
        logger.warning(f"Dependencia '{nombre_dependencia_predeterminada}' no encontrada")
    
    # Opción 2: Usar la primera dependencia
    try:
        dependencia = Dependencia.objects.first()
        if dependencia:
            logger.info(f"Usando primera dependencia disponible: {dependencia.nombre}")
            return exportar_bienes_por_dependencia_excel(request, dependencia_id=dependencia.id)
    except Exception as e:
        logger.error(f"Error al obtener primera dependencia: {e}")
    
    # Opción 3: Reporte de error
    logger.error("No se pudo encontrar ninguna dependencia")
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    fecha_descarga = datetime.now().strftime("%Y%m%d_%H%M")
    response['Content-Disposition'] = f'attachment; filename="error_reporte_{fecha_descarga}.xlsx"'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Error"
    
    ws['A1'] = "ERROR AL GENERAR REPORTE"
    ws['A3'] = "No se pudo encontrar la dependencia predeterminada."
    ws['A4'] = f"Nombre buscado: '{nombre_dependencia_predeterminada}'"
    ws['A5'] = "Por favor, configure al menos una dependencia en el sistema."
    
    wb.save(response)
    return response