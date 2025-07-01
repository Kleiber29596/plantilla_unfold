import os
from django.conf                        import settings
import openpyxl
from django.http                        import HttpResponse
from io                                 import BytesIO
from apps.cuenta.models                 import User        # Importa el modelo User si aún no lo tienes
from datetime                           import date
from apps.gestion.models.director       import Director
from apps.auxiliares.models.responsable import Responsable
from openpyxl                           import load_workbook
from openpyxl.utils                     import get_column_letter
from apps.auxiliares.models.unidad      import Unidad
from apps.cuenta.models                 import User        # Importa el modelo User si aún no lo tienes
from apps.gestion.models.asignacion     import Asignacion






def reporte_general(request):
    ruta_plantilla = os.path.join(settings.MEDIA_ROOT, 'plantillas', 'formato_inventario.xlsx')
    if not os.path.exists(ruta_plantilla):
        return HttpResponse("Plantilla no encontrada.", status=404)

    wb = openpyxl.load_workbook(ruta_plantilla)
    plantilla = wb.active  # Hoja base

    # Obtener la unidad OTIC y sus divisiones
    otic = Unidad.objects.get(descripcion="OFICINA DE TECNOLOGIAS DE LA INFORMACION Y COMUNIC")
    divisiones = Unidad.objects.filter(unidad_padre=otic)
    # print(divisiones)
    
    for division in divisiones:
        bienes = Asignacion.objects.filter(ubicacion=division.pk)
        print(bienes)
        # Crear nueva hoja con nombre de la división
        ws = wb.copy_worksheet(plantilla)
        ws.title = division.descripcion[:31]  # Excel permite máximo 31 caracteres en el nombre de hoja

        fila_inicio = 30
        for i, bien in enumerate(bienes):
            fila = fila_inicio + i
            ws[f'B{fila}'] = f'{bien.bien.categoria.cod_catalogo}'
            ws[f'C{fila}'] = bien.bien.cod_bien
            ws[f'D{fila}'] = str(bien.bien)
            ws[f'E{fila}'] = bien.bien.valor_unitario
            ws[f'F{fila}'] = bien.bien.valor_unitario

        # Agregar otros datos personalizados si deseas
        ws['B26'] = f'DENOMINACIÓN: {division.descripcion}'
        # Datos del director
        ws['B44'] = f'{otic.director.persona.origen}-{otic.director.persona.cedula}'
        ws['D44'] = f'{otic.director.persona.nombres_apellidos}'
        ws['E44'] = f'{otic.director.persona.cargo}'
       #Usuario encargado del levantamiento del inventario
        usuario =request.user
    
        ws['B43'] = f'{usuario.origen}-{usuario.cedula}'
        ws['D43'] = usuario.nombre_apellido 
        fecha_hoy = date.today()
        ws['F17'] = f'FECHA: {fecha_hoy.strftime("%d/%m/%Y")}'

    # Elimina hoja de plantilla original si no quieres que se exporte vacía
    wb.remove(plantilla)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="reporte_OTIC_{date.today().strftime("%d-%m-%Y")}.xlsx"'
    return response
